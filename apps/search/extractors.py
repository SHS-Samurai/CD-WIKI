from pathlib import Path
from zipfile import BadZipFile, ZipFile

from bs4 import BeautifulSoup
from django.conf import settings
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_attachment_text(path: Path, filename: str, content_type: str = "") -> str:
    extension = Path(filename).suffix.lower()
    try:
        if extension == ".pdf" or content_type == "application/pdf":
            text = _extract_pdf(path)
        elif extension == ".docx":
            text = _extract_docx(path)
        elif extension == ".xlsx":
            text = _extract_xlsx(path)
        elif extension in {".txt", ".md"} or content_type.startswith("text/plain"):
            text = _extract_text_file(path)
        elif extension == ".html" or content_type in {"text/html", "application/xhtml+xml"}:
            text = _extract_html(path)
        else:
            text = ""
    except Exception:
        return ""
    return _limit_text(text)


def _extract_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    if len(reader.pages) > settings.WIKI_MAX_PDF_PAGES:
        raise ValueError("PDF hat zu viele Seiten.")
    return _join_limited(page.extract_text() or "" for page in reader.pages)


def _extract_docx(path: Path) -> str:
    _validate_archive_limits(path)
    document = Document(str(path))
    return _join_limited(paragraph.text for paragraph in document.paragraphs)


def _extract_xlsx(path: Path) -> str:
    _validate_archive_limits(path)
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    parts: list[str] = []
    text_length = 0
    cell_count = 0
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    cell_count += 1
                    if cell_count > settings.WIKI_MAX_SPREADSHEET_CELLS:
                        return "\n".join(parts)
                    text_length, complete = _append_limited(parts, str(value), text_length)
                    if not complete:
                        return "\n".join(parts)
    finally:
        workbook.close()
    return "\n".join(parts)


def _extract_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_html(path: Path) -> str:
    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="replace"), "lxml")
    for tag in soup(["script", "style"]):
        tag.decompose()
    return soup.get_text(" ", strip=True)


def _limit_text(text: str) -> str:
    limit = settings.WIKI_SEARCH_ATTACHMENT_TEXT_LIMIT
    if len(text) <= limit:
        return text
    return text[:limit]


def _join_limited(values) -> str:
    parts: list[str] = []
    text_length = 0
    for value in values:
        text_length, complete = _append_limited(parts, value, text_length)
        if not complete:
            break
    return "\n".join(parts)


def _append_limited(parts: list[str], value: str, text_length: int) -> tuple[int, bool]:
    separator_length = 1 if parts else 0
    remaining = settings.WIKI_SEARCH_ATTACHMENT_TEXT_LIMIT - text_length - separator_length
    if remaining <= 0:
        return text_length, False
    fragment = value[:remaining]
    parts.append(fragment)
    new_length = text_length + separator_length + len(fragment)
    return new_length, len(fragment) == len(value)


def _validate_archive_limits(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
    except BadZipFile as exc:
        raise ValueError("Ungueltiges Office-Archiv.") from exc
    if len(entries) > settings.WIKI_MAX_ARCHIVE_MEMBERS:
        raise ValueError("Office-Archiv hat zu viele Eintraege.")
    if sum(entry.file_size for entry in entries) > settings.WIKI_MAX_ARCHIVE_UNCOMPRESSED_SIZE:
        raise ValueError("Office-Archiv ist entpackt zu gross.")
